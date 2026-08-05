"""Read researcher names and departments from the AI Lund member workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import TypedDict

from openpyxl import load_workbook


class Researcher(TypedDict):
    name: str
    department: str
    inferred: str


def _normalise_header(value: object) -> str:
    return str(value or "").strip().casefold()


def _guess_department_from_email(email: object) -> str:
    """Derive a department-like value from an email domain."""

    address = str(email or "").strip().casefold()
    _, separator, domain = address.rpartition("@")
    if not separator:
        return ""

    domain = domain.strip().rstrip(".")
    if domain.endswith(".lu.se"):
        return domain.removesuffix(".lu.se")
    if domain.endswith(".se"):
        return domain.removesuffix(".se")
    return domain


def read_researchers(
    spreadsheet: str | Path,
    *,
    lu_only: bool = False,
) -> list[Researcher]:
    """Return researcher/department records from the first worksheet.

    The workbook's ``Organisation`` column is returned unchanged as the
    department. A separate ``inferred`` value is derived from the email domain.
    Rows with no researcher name are ignored. Set ``lu_only`` to include only
    rows marked with an ``x`` in the ``LU`` column.
    """

    path = Path(spreadsheet).expanduser()
    if not path.is_file():
        raise FileNotFoundError(f"Spreadsheet not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []

        headers = {
            _normalise_header(value): index
            for index, value in enumerate(header_row)
            if value is not None
        }

        try:
            name_index = headers["name"]
            department_index = headers["organisation"]
            email_index = headers["email"]
        except KeyError as error:
            missing = error.args[0]
            raise ValueError(f"Required column is missing: {missing}") from error

        lu_index = headers.get("lu")
        if lu_only and lu_index is None:
            raise ValueError("Cannot filter LU members: the LU column is missing")

        researchers: list[Researcher] = []
        for row in rows:
            name = str(row[name_index] or "").strip()
            if not name:
                continue

            if lu_only:
                lu_marker = str(row[lu_index] or "").strip().casefold()  # type: ignore[index]
                if lu_marker != "x":
                    continue

            department = str(row[department_index] or "").strip()
            inferred = _guess_department_from_email(row[email_index])
            researchers.append(
                {
                    "name": name,
                    "department": department,
                    "inferred": inferred,
                }
            )

        return researchers
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Read researcher names and departments from an Excel workbook."
    )
    parser.add_argument("spreadsheet", type=Path, help="Path to the .xlsx file")
    parser.add_argument(
        "--lu-only",
        action="store_true",
        help="Only include rows marked with x in the LU column",
    )
    args = parser.parse_args()

    researchers = read_researchers(args.spreadsheet, lu_only=args.lu_only)
    print(json.dumps(researchers, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
